"""Excel export service — 일일 리포트 생성."""
import io
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from services.signal_service import compute_rms

# 공통 스타일
_FONT = Font(name="맑은 고딕", size=9)
_FONT_BOLD = Font(name="맑은 고딕", bold=True, size=9)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_FILL_TITLE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FILL_GAP = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_FILL_VIB_TITLE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
_FILL_VIB_HEADER = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def generate_daily_report(cycles: list[dict], date: str) -> io.BytesIO:
    """일일 리포트 Excel 생성. BytesIO 반환."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    date_display = _format_date(date)

    _create_timeline_sheet(wb, f"Timeline_{date}", cycles, date_display)
    for node in ["R1", "R2", "R3", "R4"]:
        _create_vib_sheet(wb, f"Vib_{node}", cycles, node)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _format_date(date: str) -> str:
    """YYMMDD → YYYY-MM-DD."""
    try:
        y, m, d = int("20" + date[:2]), int(date[2:4]), int(date[4:6])
        return f"{y}-{m:02d}-{d:02d}"
    except Exception:
        return date


# ---------------------------------------------------------------------------
# Timeline sheet
# ---------------------------------------------------------------------------

def _create_timeline_sheet(wb: Workbook, name: str, cycles: list[dict], date_display: str):
    ws = wb.create_sheet(name)

    # Title
    ws["A1"] = f"{date_display} Timeline"
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=11)
    ws["A1"].fill = _FILL_TITLE

    # Headers
    for col, h in enumerate(["Time", "R1", "R2", "R3", "R4"], 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = _FONT_BOLD
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_CENTER

    # Group by minute
    time_grouped: dict[str, dict[str, list[int]]] = defaultdict(
        lambda: {"R1": [], "R2": [], "R3": [], "R4": []}
    )
    for cy in cycles:
        ts = datetime.fromisoformat(cy["timestamp"])
        key = f"{ts.hour:02d}:{ts.minute:02d}"
        device_name = cy.get("device_name", "")
        if device_name in time_grouped[key]:
            time_grouped[key][device_name].append(round(cy["mpm_mean"]))

    sorted_keys = sorted(time_grouped.keys())
    if not sorted_keys:
        return

    # Detect gaps
    gaps = _detect_gaps(sorted_keys, time_grouped)

    # Write rows
    row = 3
    idx = 0
    while idx < len(sorted_keys):
        gap = next((g for g in gaps if g["start"] == idx), None)

        if gap and not gap.get("jump"):
            # Gap indicator row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            c = ws.cell(row=row, column=1, value=f"Gap ({gap['dur']} min)")
            c.font = _FONT_BOLD
            c.fill = _FILL_GAP
            c.alignment = _ALIGN_CENTER
            idx = gap["end"] + 1
            row += 1
        else:
            _write_time_row(ws, row, sorted_keys[idx], time_grouped)
            row += 1
            if gap and gap.get("jump"):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                c = ws.cell(row=row, column=1, value=f"Gap ({gap['dur']} min)")
                c.font = _FONT_BOLD
                c.fill = _FILL_GAP
                c.alignment = _ALIGN_CENTER
                row += 1
            idx += 1

    _auto_width(ws)


def _write_time_row(ws, row, time_key, time_grouped):
    ws.cell(row=row, column=1, value=time_key).font = _FONT
    for col, node in enumerate(["R1", "R2", "R3", "R4"], 2):
        vals = time_grouped[time_key][node]
        v = round(sum(vals) / len(vals)) if vals else ""
        c = ws.cell(row=row, column=col, value=v)
        c.font = _FONT
        c.alignment = _ALIGN_CENTER


def _detect_gaps(sorted_keys, time_grouped):
    """빈 구간과 시간 점프 감지."""
    gaps = []
    gap_start = None

    for i, key in enumerate(sorted_keys):
        has_data = any(len(time_grouped[key][n]) > 0 for n in ["R1", "R2", "R3", "R4"])
        if not has_data:
            if gap_start is None:
                gap_start = i
        else:
            if gap_start is not None:
                dur = _time_diff_min(sorted_keys[gap_start], sorted_keys[i - 1]) + 1
                gaps.append({"start": gap_start, "end": i - 1, "dur": dur})
                gap_start = None

    if gap_start is not None:
        dur = _time_diff_min(sorted_keys[gap_start], sorted_keys[-1]) + 1
        gaps.append({"start": gap_start, "end": len(sorted_keys) - 1, "dur": dur})

    # Time jumps between data rows
    for i in range(len(sorted_keys) - 1):
        in_gap = any(g["start"] <= i <= g["end"] or g["start"] <= i + 1 <= g["end"] for g in gaps)
        if in_gap:
            continue
        diff = _time_diff_min(sorted_keys[i], sorted_keys[i + 1])
        if diff > 1:
            gaps.append({"start": i, "end": i, "dur": diff - 1, "jump": True})

    gaps.sort(key=lambda g: g["start"])
    return gaps


def _time_diff_min(t1: str, t2: str) -> int:
    h1, m1 = map(int, t1.split(":"))
    h2, m2 = map(int, t2.split(":"))
    return (h2 * 60 + m2) - (h1 * 60 + m1)


# ---------------------------------------------------------------------------
# Vib sheet
# ---------------------------------------------------------------------------

def _create_vib_sheet(wb: Workbook, name: str, cycles: list[dict], node: str):
    ws = wb.create_sheet(name)

    # Title
    ws["A1"] = f"{node} Vibration Analysis"
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=14)
    ws["A1"].fill = _FILL_VIB_TITLE

    headers = ["Date", "Time", "Cycle", "RPM", "PX_RMS", "PY_RMS", "PZ_RMS", "VX_RMS", "VZ_RMS"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="맑은 고딕", bold=True, size=11)
        c.fill = _FILL_VIB_HEADER
        c.alignment = _ALIGN_CENTER

    node_cycles = [c for c in cycles if c.get("device_name") == node]
    rpm_stats: dict[int, list[float]] = defaultdict(list)

    row = 3
    for cy in node_cycles:
        ts = datetime.fromisoformat(cy["timestamp"])
        rpm = cy["rpm_mean"]

        # RMS 계산 (배열 데이터가 있으면 직접 계산)
        px_rms = compute_rms(cy.get("pulse_accel_x", []))
        py_rms = compute_rms(cy.get("pulse_accel_y", []))
        pz_rms = compute_rms(cy.get("pulse_accel_z", []))
        vx_rms = compute_rms(cy.get("vib_accel_x", []))
        vz_rms = compute_rms(cy.get("vib_accel_z", []))

        vals = [
            ts.strftime("%Y-%m-%d"), ts.strftime("%H:%M:%S"),
            cy.get("cycle_index", 0), round(rpm, 2),
            round(px_rms, 4), round(py_rms, 4), round(pz_rms, 4),
            round(vx_rms, 4), round(vz_rms, 4),
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = _FONT
            c.alignment = _ALIGN_CENTER

        rpm_stats[int(rpm // 10) * 10].append(rpm)
        row += 1

    # RPM statistics section
    row += 2
    ws.cell(row=row, column=1, value="RPM Statistics").font = Font(name="맑은 고딕", bold=True, size=11)
    ws.cell(row=row, column=1).fill = _FILL_TITLE
    row += 1

    for col, h in enumerate(["RPM Range", "Count", "Min", "Max", "Average"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _FONT_BOLD
        c.fill = _FILL_TITLE
        c.alignment = _ALIGN_CENTER
    row += 1

    for rng in sorted(rpm_stats):
        vals = rpm_stats[rng]
        for col, v in enumerate([f"{rng}-{rng + 10}", len(vals), round(min(vals), 2), round(max(vals), 2), round(sum(vals) / len(vals), 2)], 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = _FONT
            c.alignment = _ALIGN_CENTER
        row += 1

    _auto_width(ws)


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3
